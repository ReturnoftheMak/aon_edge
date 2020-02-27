# %% Package Imports

import pandas as pd
import openpyxl
from sql_connection import sql_connection


# %% Variables defined here are unique to Aon Edge

mappings = {'claim':r'\\svrtcs04\Syndicate Data\Actuarial\Pricing\2_Account_Pricing\NFS_Edge\Knowledge\Data_Received\Monthly\_ColumnMapping\claim_WITHACTIONS.xlsx',
            'risk':r'\\svrtcs04\Syndicate Data\Actuarial\Pricing\2_Account_Pricing\NFS_Edge\Knowledge\Data_Received\Monthly\_ColumnMapping\risk_WITHACTIONS_V2.xlsx',
            'premium':r'\\svrtcs04\Syndicate Data\Actuarial\Pricing\2_Account_Pricing\NFS_Edge\Knowledge\Data_Received\Monthly\_ColumnMapping\premium_WITHACTIONS.xlsx'}

header_dict = {'claim':2,'risk':['Policy Nbr', 'Policy Number'],'premium':0}

id_dict = {'claim':'ID_Claim', 'risk':'ID_Policy_UNCLEANSED', 'premium':'ID_PolicyStem'}

sheet_dict = {'claim':False, 'risk':True, 'premium':False}


# %% Lets try to make use of inheritance here

class BdxCleaner(object):
    """Base Class for cleaning bordereaux for AON Edge.
    Takes multiple inputs from dictionaries detailing mapping files, headers columns,
    table IDs, and sheets to take from each type of bordereaux.
    """
    def __init__(self, bdx_file, mappings, headers, IDs, sheet_dict, bdx_type):
        self.file = bdx_file
        self.mappings = mappings
        self.headers = headers
        self.IDs = IDs
        self.sheets = sheet_dict
        self.bdx_type = bdx_type
        self.xl_file = openpyxl.load_workbook(self.file, read_only=True, data_only=True)
        self.dataframe = self.basic_cleaning()


    def get_mapping(self):
        """Get a dictionary of column mapping, and a list of the column subset 
        that we want to retain.

        Args:
            None

        Returns:
            dictionary of mappings to pass to a df.rename() function in basic_cleaning()
        """

        # Read in the mapping file
        df_mapping = pd.read_excel(self.mappings[self.bdx_type])

        keys = df_mapping.ColumnNames
        values = df_mapping.Rename

        mapping_dict = dict(zip(keys, values))

        return mapping_dict


    def find_header_row(self, sheet):
        """If the header is always the same place, the dict passed in will have an integer,
        else a list of column headers to look for at the top of the sheet and this function
        will return the header row for the given sheet.

        Args:
            sheet (str): sheetname from openpyxl workbook object to loop through
        
        Returns:
            integer to pass to header kwarg in pd,read_excel()
        """

        if self.headers[self.bdx_type] is int:
            return self.headers[self.bdx_type]
        else:
            worksheet = self.xl_file[sheet]

            for row in worksheet.iter_rows(max_row=20):
                for cell in row:
                    if cell.value in self.headers[self.bdx_type]:
                        return cell.row-1


    def basic_cleaning(self):
        """Cleaning steps to be applied to any bordereaux, utilising mappings.
        Reads in the excel file, taking the first sheet only, with a header row based on type of bdx.
        Then utilises the column mapping to rename the columns and drop any columns which were not mapped.
        Lastly the function drops any rows which don't have a value in the relevant ID column 

        For a value of true in self.sheets, the function gets all the sheets in the file, else only 1st

        Args:
            None

        Returns:
            pandas.core.frame.DataFrame with required bordereaux data
        """

        if self.sheets[self.bdx_type]:

            df_list = []

            sheet_names = [sheet for sheet in self.xl_file.sheetnames if 'taxes' not in sheet.lower()]
            sheet_names = [sheet for sheet in sheet_names if 'pivot' not in sheet.lower()]

            for sheet in sheet_names:

                # Read the excel in with specified vars
                df_i = pd.read_excel(self.file, sheet_name=sheet, header=self.find_header_row(sheet))

                # Map cols using the dictionary
                df_i = df_i.rename(columns=self.get_mapping())
                df_i = df_i[[col for col in df_i.columns if type(col) is not float]]

                # Drop any subtotal rows
                df_i.dropna(axis=0, how='any', subset=[self.IDs[self.bdx_type]], inplace=True)

                df_list.append(df_i)

                df = pd.concat(df_list, ignore_index=True)

        else:
            # Read the excel in with specified vars
            df = pd.read_excel(self.file, sheet_name=0, header=self.find_header_row(sheet))

            # Map cols using the dictionary
            df = df.rename(columns=self.get_mapping())
            df = df[[col for col in df.columns if type(col) is not float]]

            # Drop any subtotal rows
            df.dropna(axis=0, how='any', subset=[self.IDs[self.bdx_type]], inplace=True)

        return df


    def username_input(self):
        """Adds in a username for whoever ran the code.

        Args:
            None
        
        Returns:
            None
        """

        from getpass import getuser

        user = getuser()

        self.dataframe['Updated_Name'] = user


    def date_code_run(self):
        """Adding in a run date for the df.

        Args:
            None
        
        Returns:
            None
        """

        from datetime import date

        today = date.today()

        self.dataframe['Updated_Date'] = today


    def add_file_name(self):
        """Add in the name of the file and the bdx month into the dataframe as columns.

        Args:
            None
        
        Returns:
            None
        """

        from pathlib import Path

        file_name = Path(self.file).stem
        bdx_month = Path(self.file).parent.parent.stem

        self.dataframe['Updated_Source'] = file_name
        self.dataframe['Bordereau_Month'] = bdx_month


    def export_to_sql(self, df, server_name, database_name, table_name, schema='bdx'):
        """Exports the cleaned dataframe to sql, replacing the old version if existing.

        Args:
            df (pandas.core.series.DataFrame): self.dataframe to export
            server_name (str): name of the SQL server (eg. tcspmSMDB02)
            database_name (str): name of the database (eg. PricingDevelopment)
            table_name (str): name of the table to export to
            schema (str): name of the database schema
        
        Returns:
            None
        """

        sql_con = sql_connection(server_name, database_name)

        df.to_sql(table_name, sql_con, schema=schema, if_exists='replace')

